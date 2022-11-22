import pandas as pd
from bs4 import BeautifulSoup  # 파싱된 데이터를 python에서 사용하기 좋게 변환
import re
import time
import warnings
import back_data_mine
from selenium import webdriver  # webdriver를 통해 파싱하기 위함
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
import zipfile
import glob
import os
from openpyxl import load_workbook

#첫번째 시도에 문제 없는 OK건 다 올린다. (note = 'OK', 구매수량 >0)
#두번째 시도에 수정, 추가건은 파일 생성까지만 된다. (자동으로 딜리버드에 올라가지 않는다.)

#배송메시지 명령어 : skip <- 수량 0으로 처리 (오류주문, 주문 안되게), hold <- (반품 기다리는 것) 주문 안들어 가게..
#링크 없는건 미리 배송메시지에 업데이트 해놓는다.

add = 1

if add == 0:
    print("전체작업 (첫번째 시도, 정상상품은 딜리버드 등록까지)")
else:
    print("추가작업:",add,"개")

# 기본세팅
warnings.filterwarnings("ignore")

options = webdriver.ChromeOptions()
options.headless = True
options.add_argument("window-size=1920x1080")
search_dir = "/Users/seoyulejo/Downloads/shopping_raw/" # 필요한 경우 편집
prefs = {'download.default_directory' : search_dir}
options.add_experimental_option('prefs', prefs)
options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36")
driver = webdriver.Chrome("/Users/seoyulejo/chromedriver", options=options) #, options=options
driver.maximize_window()
driver.implicitly_wait(15)
action = ActionChains(driver)
wait = WebDriverWait(driver, 20)

category_list = back_data_mine.category_list # 분류설정

# 신상마켓 로그인
driver.get('https://sinsangmarket.kr/login')
"""try:
    driver.find_element_by_xpath('//*[@id="alert"]/div/div/button').click() #too many segment 버튼 클릭
except:
    pass
"""
try:
    driver.find_element_by_xpath('//*[@id="app"]/div[1]/div/header/div/div[2]/div[3]/p').click() #start 버튼 클릭
except:
    pass

time.sleep(.5)
driver.find_element_by_xpath('//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[1]/input').click()
action.send_keys('protestt').perform()
time.sleep(.5)
driver.find_element_by_xpath('//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[2]/input').click()
action.send_keys('!QAZwsx123').perform()
time.sleep(.5)
driver.find_element_by_xpath('//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[3]/div[2]/div/button').click()
print("신상 로그인 성공")

# 광고 있으면 close
time.sleep(.5)
try:
    driver.find_element_by_class_name("button.close-button").click()
    time.sleep(.3)
except:
    pass

# 한글로 바꾸기
driver.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[1]/div/ul/li[5]/div/div').click()
time.sleep(.5)
driver.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[1]/div/ul/li[5]/div/ul/li[1]/label').click()
time.sleep(.5)

# 광고 있으면 close
try:
    driver.find_element_by_class_name("button.close-button").click()
    time.sleep(.3)
except:
    pass

# cafe24 열기
driver.switch_to.new_window('tab')
time.sleep(.5)
driver.switch_to.window(driver.window_handles[1])
driver.get("https://eclogin.cafe24.com/Shop/")
time.sleep(.5)
driver.find_element_by_xpath('//*[@id="mall_id"]').click()
time.sleep(.5)
action.send_keys('soyool').perform()
driver.find_element_by_xpath('//*[@id="userpasswd"]').click()
action.send_keys('!QAZwsx123').perform()
time.sleep(.5)
driver.find_element_by_xpath('//*[@id="frm_user"]/div/div[3]/button').click()
time.sleep(3)
print("cafe24 진입")

driver.get('https://soyool.cafe24.com/admin/php/shop1/s_new/shipped_begin_list.php') #배송준비중으로 이동
time.sleep(2)
driver.find_element_by_xpath('//*[@id="QA_deposit1"]/div[2]/table/tbody/tr[2]/td/a[6]').click() #지난 한달 선택
time.sleep(1)

select = Select(driver.find_element_by_xpath('//*[@id="QA_prepareNumber2"]/div[5]/div[2]/select[2]'))  # 검색종류
select.select_by_visible_text('100개씩보기')
time.sleep(1)

driver.find_element_by_xpath('//*[@id="search_button"]').click() #검색

try:
    driver.find_element_by_xpath('//*[@id="ch-plugin-core"]/div[2]/div/div/button').click()  # 광고 close
except:
    pass

#엑셀 다운로드
driver.find_element_by_xpath('//*[@id="eExcelDownloadBtn"]').click()
time.sleep(1)
driver.switch_to.window(driver.window_handles[2])

select = Select(driver.find_element_by_xpath('//*[@id="aManagesList"]'))
select.select_by_visible_text('ForOrderSejo')
time.sleep(.5)

driver.find_element_by_xpath('//*[@id="Password"]').click()
action.send_keys('protest123').perform()
time.sleep(.5)

driver.find_element_by_xpath('//*[@id="PasswordConfirm"]').click()
action.send_keys('protest123').perform()
time.sleep(.5)

driver.find_element_by_xpath('//*[@id="QA_common_password1"]/div[3]/a').click() #자료 요청
time.sleep(.5)

alert = driver.switch_to.alert
alert.accept()
driver.switch_to.window(driver.window_handles[2])

time.sleep(10)
driver.find_element_by_xpath('//*[@id="QA_common_password2"]/div[3]/table/tbody/tr[1]/td[7]').click() #다운로드 버튼

driver.find_element_by_xpath('//*[@id="password"]').click() #다시 비밀번호 물어봄
action.send_keys('protest123').perform()
time.sleep(.5)

driver.find_element_by_xpath('//*[@id="reason_for_download"]').click() #이유 적어줌
action.send_keys('analysis').perform()
time.sleep(.5)

driver.find_element_by_xpath('//*[@id="excel_download"]').click() #파일요청
time.sleep(7)
driver.close()
driver.switch_to.window(driver.window_handles[1])

print("다운로드 완료")
#다운로드 zip 파일 이름 확인 https://stackoverflow.com/questions/168409/how-do-you-get-a-directory-listing-sorted-by-creation-date-in-python
files = list(filter(os.path.isfile, glob.glob(search_dir + "*")))
files.sort(key=lambda x: os.path.getmtime(x))
file_path = files[-1]

#zip 풀기
with zipfile.ZipFile(file_path, 'r') as zip_ref:
    zip_ref.extractall(search_dir, pwd=b'protest123')
print("압축해제 완료")

#변환 파일 이름 확인 https://stackoverflow.com/questions/168409/how-do-you-get-a-directory-listing-sorted-by-creation-date-in-python
files = list(filter(os.path.isfile, glob.glob(search_dir + "*")))
files.sort(key=lambda x: os.path.getmtime(x))
file_name = files[-1]

df = pd.read_csv(file_name)
print("df import 완료:",len(df),"개")

df['option1'] = df['상품옵션'].replace('.*=(\S+),.*',r'\1', regex=True)
df['option2'] = df['상품옵션'].replace('.*=.*=(.*)',r'\1', regex=True)
df['option2'] = df['option2'].str.lower()
df['수령인 우편번호'] = df['수령인 우편번호'].astype(str)
df['배송메시지'] = df['배송메시지'].astype(str)
zip_code = []
for i in range(len(df)):
    if len(df['수령인 우편번호'][i])==4:
        z_code = '0'+df['수령인 우편번호'][i]
    else:
        z_code = df['수령인 우편번호'][i]
    zip_code.append(z_code)
df['수령인 우편번호'] = zip_code

title_ss = []
price_ss = []
shop_name = []
building_name = []
shop_location = []
shop_phone_number = []
note = []

print("데이터 채우기 시작")

skip_point = len(df)-add
for i in range(len(df)):
    if pd.isnull(df['모델명'][i]):
        link = df['배송메시지'][i]
        df['모델명'][i] = df['배송메시지'][i]
        if df['주문경로'][i]=='쿠팡':
            option_ = df['상품옵션'][i]
            option_.strip()
            df['option1'][i] = re.search(r'옵션=(\S*)\s(\S*)', option_).group(1)
            df['option2'][i] = re.search(r'옵션=(\S*)\s(\S*)', option_).group(2)
    else:
        link = df['모델명'][i]

    if add !=0 and i < skip_point:
        note.append('이미 처리')
        title_ss.append('-')
        price_ss.append('0')
        shop_name.append('-')
        building_name.append('-')
        shop_location.append('-')
        shop_phone_number.append('-')
        print(i,"이미등록")
        continue

    driver.switch_to.new_window('tab')
    time.sleep(.5)
    driver.switch_to.window(driver.window_handles[2])
    #url 검
    try:
        driver.get(link)
        time.sleep(1.5)
    except:
        note.append('!! url 없음 !!')
        title_ss.append('!! url 없음 !!')
        price_ss.append('!! url 없음 !!')
        shop_name.append('!! url 없음 !!')
        building_name.append('!! url 없음 !!')
        shop_location.append('!! url 없음 !!')
        shop_phone_number.append('!! url 없음 !!')

        driver.close()
        driver.switch_to.window(driver.window_handles[1])
        time.sleep(.5)
        print(i+1,"url 없음 skip")
        continue

    # 도매상품이름
    try:
        title = driver.find_element_by_xpath('//*[@id="goods-detail"]/div/div[2]/div[2]/div[1]/p').text
        title_ss.append(title)
    except:
        note.append('!! link 이상 !!')
        title_ss.append('!! link 이상 !!')
        price_ss.append('!! link 이상 !!')
        shop_name.append('!! link 이상 !!')
        building_name.append('!! link 이상 !!')
        shop_location.append('!! link 이상 !!')
        shop_phone_number.append('!! link 이상 !!')

        driver.close()
        driver.switch_to.window(driver.window_handles[1])
        time.sleep(.5)
        print(i + 1, "link 이상 skip")
        continue

    #품절확인
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    if soup.find("div", attrs={'class': 'sold-out'}):
        note.append('OK - sold-out')
    else:
        note.append('OK')

    time.sleep(.5)
    #도매가격
    price = int(driver.find_element_by_xpath('//*[@id="goods-detail"]/div/div[2]/div[2]/div[1]/div[3]/div[1]/div/span').text.replace(",",""))
    price_ss.append(price)
    time.sleep(.5)

    #가게 화면 진입
    driver.find_element_by_xpath('//*[@id="goods-detail"]/div/div[2]/div[2]/div[1]/div[1]/span').click()
    time.sleep(1.5)
    driver.switch_to.window(driver.window_handles[3])
    time.sleep(.5)
    #가게이름
    shop_n = driver.find_element_by_xpath('/html/body/div/div[1]/div[2]/div/div/div[1]/div[2]/div[1]/div[1]/span').text
    shop_name.append(shop_n)
    time.sleep(.5)
    #전화번호
    shop_p = driver.find_element_by_xpath('/html/body/div/div[1]/div[2]/div/div/div[1]/div[2]/div[1]/div[2]/div[1]/div').text.strip()
    shop_phone_number.append(shop_p)
    time.sleep(.5)

    address = driver.find_element_by_xpath('/html/body/div/div[1]/div[2]/div/div/div[1]/div[2]/div[1]/div[2]/div[2]/div').text.strip()
    #빌딩이름
    try:
        building_n = re.search(r'(.*)\s\S+층\s.*', address).group(1)
    except:
        try:
            building_n = re.search(r'(.*\s별관)\s.*', address).group(1)
        except:
            building_n = address
    building_name.append(building_n)
    time.sleep(.5)
    #주소
    try:
        shop_l = re.search(r'.*\s(\S+층\s.*)', address).group(1)
    except:
        try:
            shop_l = re.search(r'.*\s별관\s(.*)', address).group(1)
        except:
            shop_l = address
    shop_location.append(shop_l)
    time.sleep(.5)

    driver.close()
    driver.switch_to.window(driver.window_handles[2])
    driver.close()
    driver.switch_to.window(driver.window_handles[1])
    print(i+1,"완료")

print("데이터 채우기 완료")

df['title_ss'] = title_ss
df['price_ss'] = price_ss
df['shop_name'] = shop_name
df['building_name'] = building_name
df['shop_location'] = shop_location
df['shop_phone_number'] = shop_phone_number
df['note'] = note

cols = df.columns.tolist()
cols = cols[:11]+cols[12:14]+cols[15:19]+cols[11:12]+cols[14:15]+cols[19:]
df = df[cols]

df.insert(12,'key','')
df.insert(15,'in_stock','')
df.insert(16,'구매수량','')
df.insert(17,'미송수량','')
df.insert(18,'실구매수량','')
df.insert(19,'미송노트','')

#재고와 비교위한 key값 생성
df['key'] = df['상품명(한국어 쇼핑몰)']+"_"+df['option1']+"_"+df['option2']
df['key'] = df['key'].str.lower()
df['key'] = df['key'].replace('\s','', regex=True)

####################
#딜리버드에서 재고 다운받기
driver.switch_to.window(driver.window_handles[0])
#딜리버드로 가기
driver.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[1]/div/ul/li[1]/div').click()
time.sleep(2)

print("딜리버드 진입 - 반품재고 확인 (상품 및 재고 탭)")
#재고로 가기
driver.find_element_by_xpath('//*[@id="navbarSupportedContent"]/ul/li[7]/a').send_keys(Keys.ENTER)
time.sleep(3)
driver.find_element_by_xpath('//*[@id="page-wrapper"]/div[2]/div[2]/div/div/div/div[3]/div/div/div[2]/label').send_keys(Keys.SPACE) #재고 없음 클릭
time.sleep(3)
driver.find_element_by_xpath('//*[@id="productList_wrapper"]/div[1]/div[2]/div/button[1]').send_keys(Keys.ENTER)
time.sleep(4)

files = list(filter(os.path.isfile, glob.glob(search_dir + "*")))
files.sort(key=lambda x: os.path.getmtime(x))
file_name2 = files[-1]

time.sleep(1)
df_stock = pd.read_excel(file_name2)
df_stock['key'] = df_stock['판매 상품명']+"_"+df_stock['상품옵션 1']+"_"+df_stock['상품옵션 2']
df_stock['key'] = df_stock['key'].str.lower()
df_stock['key'] = df_stock['key'].replace('\s','', regex=True)

#반품 상품중 옵션 다른 것 팔렸을때 표시, 구매 안함.
df_stock_2 = df_stock[['판매 상품명','상품옵션 1','상품옵션 2']]
list_2 = df_stock_2.values.tolist()
for i in range(len(df)):
    for j in range(len(df_stock_2)):
        if df['상품명(한국어 쇼핑몰)'][i] == list_2[j][0]:
            if df['option1'][i] != list_2[j][1] or df['option2'][i] != list_2[j][2]:
                df['note'][i] = "옵션 다른 반품 상품, 취소해야함."

#과거 반품상품 점검
for i in range(len(df)):
    if df['상품명(한국어 쇼핑몰)'][i] in back_data_mine.block_subject:
        df['note'][i] = "과거 반품상품과 같은이름, 점검 필요."

df_stock_ = df_stock[['key','상품번호','정상재고']]
list_ = df_stock_.values.tolist()
dict_ = {}
for i in range(len(list_)):
    dict_[list_[i][0]] = [list_[i][1],list_[i][2]]
print("재고 dic 완료")

# 반품 재고 매입 수량에 반영..
p_result = [] # 사입번호
p_number = [] # stock 수량
for i in range(len(df)):
    result = '' #번호
    num = 0 #수량
    if df['배송메시지'][i].lower() == "skip":
        df['수량'][i] = 0
    if df['배송메시지'][i].lower() == "hold":
        df['note'][i] = "hold"
    if df['key'][i].lower() in dict_:
        # 사입번호 넣기
        result = dict_[df['key'][i].lower()][0]
        if dict_[df['key'][i].lower()][1]>0:
            #stock 개수 넣기
            for j in range(df['수량'][i].item()):
                if dict_[df['key'][i].lower()][1]>0: #재고개수 >0?
                    num +=1
                    dict_[df['key'][i].lower()][1] -= 1
                    if dict_[df['key'][i].lower()][1] ==0:
                        df['note'][i] = "재고 소진, 상품삭제 필요"
    p_result.append(result)
    p_number.append(num)

df['temp_사입번호'] = p_result
df['in_stock'] = p_number
df['구매수량'] = df['수량']-df['in_stock']
print("df에 재고수량 반영")

#미송 마스터에 반영
print("딜리버드 진입 - 미송 확인 (사입현황 탭)")
driver.find_element_by_xpath('//*[@id="navbarSupportedContent"]/ul/li[2]/a').send_keys(Keys.ENTER) #재고로 가기
time.sleep(2)
driver.find_element_by_xpath('//*[@id="page-wrapper"]/div[2]/div[1]/div[1]/ul/li[3]/a').send_keys(Keys.ENTER) #미송상품 클릭
time.sleep(3)
driver.find_element_by_xpath('//*[@id="prepaidProduct_wrapper"]/div[1]/div[2]/div/button').send_keys(Keys.ENTER) # 엑셀 다운로
time.sleep(4)

files = list(filter(os.path.isfile, glob.glob(search_dir + "*")))
files.sort(key=lambda x: os.path.getmtime(x))
file_name3 = files[-1]

time.sleep(1)
df_delay = pd.read_excel(file_name3)
df_delay['미송대기수량'] = df_delay['사입 요청 수량']-df_delay['누적 사입 수량']-df_delay['환불 수량']
df_delay['상품옵션'] = df_delay['옵션 1/2'].str.replace('/','_')

df_delay['key'] = df_delay['판매 상품명']+"_"+df_delay['상품옵션']
df_delay['key'] = df_delay['key'].str.lower()
df_delay['key'] = df_delay['key'].replace('\s','', regex=True)

df_delay_ = df_delay[['key','미송대기수량','사입사 메모']]
list_d = df_delay_.values.tolist()
dict_d = {}
for i in range(len(list_d)):
    dict_d[list_d[i][0]] = [list_d[i][1],list_d[i][2]]
print("미송 dic 완료")

# 미송 - 매입 수량에서 제외..
p_number = [] # 미송 수량
note_misong = [] # 미송 노트
for i in range(len(df)):
    num = 0 #수량
    m_note =""
    if df['key'][i] in dict_d:
        if dict_d[df['key'][i]][0]>0:
            #stock 개수 넣기
            for j in range(df['구매수량'][i].item()):
                if dict_d[df['key'][i]][0]>0: #재고개수 >0?
                    num +=1
                    dict_d[df['key'][i]][0] -= 1
            m_note = dict_d[df['key'][i]][1]
    p_number.append(num)
    note_misong.append(m_note)
df['미송수량'] = p_number
df['실구매수량'] = df['구매수량']-df['미송수량']
df['미송노트'] = note_misong
print("df에 미송수량 반영")

timestr_now = time.strftime("%Y%m%d-%H%M%S")
timestr = time.strftime("%Y%m%d")
file_name_master = "/Users/seoyulejo/Downloads/files/order_master_"+timestr+".xlsx"

if add == 0:
    df.to_excel(file_name_master)
    print("엑셀 export 완료")
else:
    ExcelWorkbook = load_workbook(file_name_master)
    writer = pd.ExcelWriter(file_name_master, engine='openpyxl')
    writer.book = ExcelWorkbook
    df.to_excel(writer, sheet_name=timestr_now)
    writer.save()
    writer.close()

try:
    os.remove(file_path)
    os.remove(file_name)
    os.remove(file_name2)
    os.remove(file_name3)
except OSError as e:
    print(e.strerror)










# 여기서 부터 정상상품 자동 업로드 위해 추가 작성한 부분

print("order form 작성 시작")

# 오더 양식으로 변경
df['option1'] = df['option1'].str.lower()
df['option2'] = df['option2'].str.lower()
df2 = df.groupby(
    ['title_ss', '상품명(한국어 쇼핑몰)', 'option1', 'option2', 'price_ss', 'shop_name', 'building_name', 'shop_location',
     'shop_phone_number', '상품품목코드', '모델명', 'note'], dropna=False).agg(
    {'수량': 'sum', 'in_stock': 'sum', '실구매수량': 'sum'})
df2 = df2.add_suffix('').reset_index()
df2 = df2[df2['실구매수량'] > 0]
df2 = df2[df2['note'].str.contains('OK')]

# 도매 상품명 중복 검증위한 key_값 생성
df2['key_'] = df2['title_ss'] + "_" + df2['option1'] + "_" + df2['option2']
df2['key_'] = df2['key_'].str.lower()
df2['key_'] = df2['key_'].replace('\s', '', regex=True)
check = []
for i in range(len(df2)):
    if df2['key_'].iloc[i] in check:
        while df2['key_'].iloc[i] in check:
            df2['title_ss'].iloc[i] += "_"
            df2['key_'].iloc[i] = df2['title_ss'].iloc[i] + "_" + df2['option1'].iloc[i] + "_" + df2['option2'].iloc[i]
        check.append(df2['key_'].iloc[i])
    else:
        check.append(df2['key_'].iloc[i])
df2.drop('key_', axis=1, inplace=True)

cols = df2.columns.tolist()
cols = cols[:5] + cols[14:15] + cols[5:14]
df2 = df2[cols]

df2.insert(0, '고객사 상품코드_temp', '')
df2.insert(1, '일시품절시_temp', '')
df2.insert(6, '기타옵션-temp', '')
df2.insert(9, '카테고리-temp', '')
df2.insert(14, 'blank1-temp', '')
df2.insert(15, 'blank2-temp', '')
df2.insert(16, 'blank3temp', '')
df2.insert(17, 'blank4-temp', '')
df2.insert(19, 'blank5-temp', '')  # 메모2, 서율샵 자리

df2['일시품절시_temp'] = '미송'

# 공백 추가
df2.loc[-1] = ""
df2.index = df2.index + 1
df2 = df2.sort_index()

df2.loc[-1] = ""
df2.index = df2.index + 1
df2 = df2.sort_index()

file_order_form = "/Users/seoyulejo/Downloads/files/order_form_" + timestr + ".xlsx"

if add == 0:
    df2.to_excel(file_order_form, index=False)
else:
    ExcelWorkbook = load_workbook(file_order_form)
    writer = pd.ExcelWriter(file_order_form, engine='openpyxl')
    writer.book = ExcelWorkbook
    df2.to_excel(writer, sheet_name=timestr_now, index=False)
    writer.save()
    writer.close()

print("엑셀 order form - export 완료")












#딜리버드에 등록
if add == 0:
    driver.switch_to.window(driver.window_handles[0])
    time.sleep(3)
    driver.find_element_by_xpath('//*[@id="navbarSupportedContent"]/ul/li[1]/a').send_keys(Keys.ENTER) # 사입요청 탭
    time.sleep(1)

    action.send_keys(Keys.PAGE_DOWN)
    element = driver.find_element_by_xpath('//*[@id="page-wrapper"]/div[2]/div[2]/div/div[2]/div/div/div/div/div[1]/div[2]/button[2]')
    action.move_to_element(element)
    element.send_keys(Keys.ENTER) # 엑셀 등록

    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="excel_file"]').send_keys(file_order_form)  # 파일선택 버튼
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="uploadExcelModal"]/div/div/div[2]/form/div[3]/input').send_keys(Keys.ENTER) # 보내기 버튼(저장)
    time.sleep(6)

    try:
        driver.find_element_by_xpath('//*[@id="purchasesList_wrapper"]/div[1]/div/div/button[9]').send_keys(Keys.ENTER)  #사입요청하기 버튼
    except:
        driver.find_element_by_xpath('//*[@id="purchasesList_wrapper"]/div[1]/div/div/button[9]').send_keys(
            Keys.ENTER)  # 사입요청하기 버튼

    element = driver.find_element_by_xpath('/html/body/div[5]/div/div[3]/button[3]')
    time.sleep(1)
    element.send_keys(Keys.ENTER)  # 네

    time.sleep(4)
    element = driver.find_element_by_xpath('//*[@id="method_SINSANGPOINT"]')
    time.sleep(.5)
    element.send_keys(Keys.ENTER) # 신상캐시 선택
    time.sleep(1)

    for i in range(8):
        action.send_keys(Keys.DOWN)
        time.sleep(1)

    element = driver.find_element_by_xpath('//*[@id="confirmCollapse"]/div[2]/div/label/span[1]') # 동의합니다.
    try:
        element.send_keys(Keys.SPACE)
    except:
        element.click()

    time.sleep(2)
    element = driver.find_element_by_xpath('//*[@id="payment_button"]')
    element.send_keys(Keys.ENTER)  # 결재버튼
    time.sleep(6)
    print("결재 완료")






